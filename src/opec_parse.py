"""Turn a downloaded MOMR (Excel appendix + PDF) into the synopsis JSON that
opecreport.py renders — figures, chart series, and deterministic prose.

Data sources (verified against the June 2026 issue):
  * Excel appendix Table 11-1  — world demand & production balance (levels by year/qtr)
  * Excel appendix Table 11-2  — changes from last month (the m/m revisions)
  * Excel appendix Table 11-4  — non-DoC liquids by country (the 26/25 contributor column)
  * PDF Oil Market Highlights  — feature article, as-of date, price narrative
  * PDF secondary-sources table — DoC crude production by country (latest month + m/m)
  * PDF Table 1-1              — OPEC Reference Basket price

The prose (headline / implications / changes) is generated from the parsed numbers
with rules, so the whole pipeline runs unattended with no API key. Every section is
best-effort: if a table is missing or the layout shifts, that piece is dropped and the
report still builds from whatever parsed. Keep all wording neutral observation — no
buy/sell/recommend language (compliance).
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pypdfium2 as pdfium

_MONTHS = ("January February March April May June July August September October "
           "November December").split()
_ABBR = {m[:3]: m for m in _MONTHS}


# ------------------------------------------------------------------ helpers
def _f(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def _rows(ws):
    return [[c for c in row] for row in ws.iter_rows(values_only=True)]


def _colmap(rows):
    """Find the header row that carries the year/quarter labels and map label->col idx."""
    for r in rows:
        cells = [str(c).strip() if c is not None else "" for c in r]
        if "2025" in cells and "2026" in cells and "2027" in cells:
            return {lbl: i for i, lbl in enumerate(cells) if lbl}
    return {}


def _find(rows, label, after=None):
    """First row whose first non-empty cell equals `label` (optionally after a marker row)."""
    started = after is None
    for r in rows:
        first = next((str(c).strip() for c in r if c is not None and str(c).strip()), "")
        if not started:
            if first == after:
                started = True
            continue
        if first.replace("\n", " ").strip() == label:
            return r
    return None


def _val(row, cmap, col):
    if row is None or col not in cmap or cmap[col] >= len(row):
        return None
    return _f(row[cmap[col]])


# ------------------------------------------------------------------ appendix
def parse_appendix(xlsx: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    out = {}
    t1 = _rows(wb["Table 11 - 1"]) if "Table 11 - 1" in wb.sheetnames else []
    cm = _colmap(t1)
    if cm:
        dem = _find(t1, "(a) Total world demand")
        oecd = _find(t1, "Total OECD")                       # first = demand-side OECD
        noecd = _find(t1, "Total Non-OECD")
        ndl = _find(t1, "Total Non-DoC liquids production")
        b = _find(t1, "(b) Total Non-DoC liquids production and DoC NGLs")
        call = _find(t1, "(a) - (b)")
        docp = _find(t1, "DoC crude oil production")
        yrs = ["2024", "2025", "2026", "2027"]

        def growth(row):
            return {y: (None if _val(row, cm, y) is None or _val(row, cm, str(int(y) - 1)) is None
                        else round(_val(row, cm, y) - _val(row, cm, str(int(y) - 1)), 2)) for y in yrs}

        dg, og, ng, sg = growth(dem), growth(oecd), growth(noecd), growth(ndl)
        out["demand_by_year"] = [{"year": y, "oecd": og[y], "nonoecd": ng[y], "total": dg[y]}
                                 for y in yrs if dg[y] is not None]
        out["supply_by_year"] = [{"year": y, "nondoc": sg[y]} for y in yrs if sg[y] is not None]
        out["call"] = {c: _val(call, cm, c) for c in ("1Q26", "2Q26", "3Q26", "4Q26", "2026", "2027", "2025")}
        out["doc_prod"] = {c: _val(docp, cm, c) for c in ("2025", "1Q26")}
        out["levels"] = {
            "demand2026": _val(dem, cm, "2026"), "demand2025": _val(dem, cm, "2025"),
            "ndl2026": _val(ndl, cm, "2026"),
        }
        bq = []
        for q in ("1Q26", "2Q26", "3Q26", "4Q26", "2026", "2027"):
            cv = _val(call, cm, q)
            if cv is None:
                continue
            row = {"q": q, "call": round(cv, 2)}
            if q == "1Q26" and _val(docp, cm, "1Q26") is not None:
                row["prod"] = round(_val(docp, cm, "1Q26"), 2)
            bq.append(row)
        out["balance_quarters"] = bq

    # Table 11-2 — m/m revisions (Y-o-y change rows under (a) and (b); (a)-(b) row)
    t2 = _rows(wb["Table 11 - 2"]) if "Table 11 - 2" in wb.sheetnames else []
    cm2 = _colmap(t2)
    if cm2:
        # the "Y-o-y change" row immediately after "(a) Total world demand"
        def yoy_after(marker):
            seen = False
            for r in t2:
                first = next((str(c).strip() for c in r if c is not None and str(c).strip()), "")
                if first == marker:
                    seen = True
                    continue
                if seen and first == "Y-o-y change":
                    return r
            return None
        dchg, schg = yoy_after("(a) Total world demand"), yoy_after("(b) Total Non-DoC liquids production and DoC NGLs")
        callchg = _find(t2, "(a) - (b)")
        out["rev"] = {
            "demand_growth_2026": _val(dchg, cm2, "2026"), "demand_growth_2027": _val(dchg, cm2, "2027"),
            "supply_growth_2026": _val(schg, cm2, "2026"), "supply_growth_2027": _val(schg, cm2, "2027"),
            "call_2026": _val(callchg, cm2, "2026"), "call_2027": _val(callchg, cm2, "2027"),
        }

    # Table 11-4 — non-DoC contributors (26/25 change column)
    t4 = _rows(wb["Table 11 - 4"]) if "Table 11 - 4" in wb.sheetnames else []
    if t4:
        # header has two "Change" cells; the 26/25 change is the col right after the "2026" col
        hdr_i = next((i for i, r in enumerate(t4)
                      if "2026" in [str(c).strip() for c in r if c is not None]), None)
        contribs = []
        if hdr_i is not None:
            hdr = [str(c).strip() if c is not None else "" for c in t4[hdr_i]]
            c2026 = hdr.index("2026")
            col2625 = c2026 + 1                       # the "Change" column following 2026
            skip = {"OECD Americas", "OECD Europe", "OECD Asia Pacific", "Total OECD", "Other Asia",
                    "Latin America", "Middle East", "Africa", "Other Eurasia", "Other Europe",
                    "Total Non-OECD", "Non-DoC production", "Processing gains",
                    "Non-DoC liquids production", "DoC NGLs",
                    "Non-DoC liquids production and DoC NGLs"}
            for r in t4[hdr_i + 1:]:
                name = next((str(c).strip() for c in r if c is not None and str(c).strip()), "")
                if not name or name in skip or name.startswith("Total") or name.startswith("Note"):
                    continue
                v = _f(r[col2625]) if col2625 < len(r) else None
                if v is not None:
                    contribs.append({"name": name.replace("others", "(other)"), "value": round(v, 2)})
        # top movers by magnitude, keep a sensible handful
        contribs.sort(key=lambda d: -abs(d["value"]))
        out["supply_contributors"] = contribs[:7]
    return out


# ------------------------------------------------------------------ pdf
def _pdf_text(pdf: Path, pages=None) -> str:
    d = pdfium.PdfDocument(str(pdf))
    idx = range(len(d)) if pages is None else pages
    return "\n".join(d[i].get_textpage().get_text_range() for i in idx if i < len(d))


def parse_pdf(pdf: Path) -> dict:
    out = {}
    head = _pdf_text(pdf, range(0, 14))
    # as-of date + feature article
    m = re.search(r"(\d{1,2}\s+(?:" + "|".join(_MONTHS) + r")\s+\d{4})", head)
    if m:
        out["asof"] = m.group(1)
    fm = re.search(r"Feature article:\s*(.+?)(?:\n\n|\nOil market highlights)", head, re.S)
    if fm:
        out["feature"] = re.sub(r"\s+", " ", fm.group(1)).strip()
    # ORB price (Table 1-1: "ORB 109.06 114.55 5.49 72.50 94.50" with header months)
    hm = re.search(r"OPEC Reference Basket \(ORB\)\s+([A-Z][a-z]{2}\s+\d{2})\s+([A-Z][a-z]{2}\s+\d{2})", head)
    orb = re.search(r"\n\s*ORB\s+([\d.]+)\s+([\d.]+)\s+(-?[\d.]+)\s+([\d.]+)\s+([\d.]+)", head)
    if orb:
        prev_m, last_m = (hm.group(1), hm.group(2)) if hm else ("prev", "latest")
        out["orb"] = {"prev_label": prev_m, "last_label": last_m,
                      "prev": _f(orb.group(1)), "last": _f(orb.group(2)), "chg": _f(orb.group(3)),
                      "y2025": _f(orb.group(4)), "y2026": _f(orb.group(5))}
    # narrative flags for prose
    full_head = head.lower()
    out["backwardated"] = "backwardat" in full_head
    out["specs_cut"] = bool(re.search(r"(cut|reduced|trimmed).{0,40}(long|bullish|net.?long)", full_head))
    # DoC crude production by country (secondary sources)
    out["doc_production"], out["doc_prod_current"] = _parse_secondary(pdf)
    return out


def _parse_secondary(pdf: Path):
    """The secondary-sources crude-production table: each country line is
    'Name n n n n n n n n change' in tb/d; take the latest month (2nd-last number)
    and the m/m change (last). The block ends at the first 'Total DoC' line."""
    txt = _pdf_text(pdf)
    # locate the block: from the 'averaged X mb/d' sentence to the first 'Total DoC'
    anchor = txt.find("Total DoC crude oil production averaged")
    seg = txt[anchor: anchor + 4000] if anchor >= 0 else txt
    end = seg.find("Total DoC")
    block = seg[:end] if end > 0 else seg
    want = ["Saudi Arabia", "Iraq", "IR Iran", "Kuwait", "UAE", "Nigeria", "Libya", "Algeria",
            "Venezuela", "Congo", "Gabon", "Equatorial Guinea", "Russia", "Kazakhstan", "Mexico",
            "Oman", "Azerbaijan", "Malaysia", "Bahrain", "Brunei", "South Sudan", "Sudan"]
    rename = {"IR Iran": "Iran"}
    prod, current = [], None
    for name in want:
        m = re.search(re.escape(name) + r"\s+((?:[\d,]+\s+){5,}[\d,]+)\s+(-?[\d,]+)", block)
        if not m:
            continue
        nums = [int(s.replace(",", "")) for s in m.group(1).split() if s.replace(",", "").isdigit()]
        chg = int(m.group(2).replace(",", "")) if m.group(2).replace(",", "").lstrip("-").isdigit() else None
        if not nums:
            continue
        latest = nums[-1]                       # last value in the run = latest month
        prod.append({"country": rename.get(name, name), "value": round(latest / 1000, 2),
                     "mom": round(chg / 1000, 2) if chg is not None else None})
    # total DoC current (May) from "Total DoC crude oil production averaged 33.13 mb/d"
    tm = re.search(r"DoC crude oil production averaged (?:about )?([\d.]+)\s*mb/d", txt)
    if tm:
        current = _f(tm.group(1))
    prod.sort(key=lambda d: -d["value"])
    return prod, current


# ------------------------------------------------------------------ prose
def _rev_phrase(v, unit="mb/d"):
    if v is None:
        return "—"
    if abs(v) < 0.05:
        return "unch."
    return f"{v:+.1f}"


def _rev_word(v):
    if v is None or abs(v) < 0.05:
        return "left broadly unchanged"
    return f"revised {'up' if v > 0 else 'down'} {abs(v):.1f} mb/d"


def build_synopsis(appendix: dict, pdf: dict, month: str, year: str) -> dict:
    d = {"month": f"{month} {year}", "verified": True,
         "asof": pdf.get("asof", ""), "feature": pdf.get("feature", "")}

    def gyear(key, y):
        for r in appendix.get(key, []):
            if r["year"] == y:
                return r
        return {}
    dem26, dem27 = gyear("demand_by_year", "2026"), gyear("demand_by_year", "2027")
    sup26 = gyear("supply_by_year", "2026")
    call = appendix.get("call", {})
    rev = appendix.get("rev", {})
    orb = pdf.get("orb", {})
    docout = pdf.get("doc_prod_current")
    latest_month = (orb.get("last_label", "") or "the latest month").split()[0] if orb else "the latest month"

    dg26 = dem26.get("total"); dg27 = dem27.get("total"); sg26 = sup26.get("nondoc")
    c26 = call.get("2026"); c27 = call.get("2027")

    # chart series straight through
    for k in ("demand_by_year", "supply_by_year", "supply_contributors", "balance_quarters"):
        if appendix.get(k):
            d[k] = appendix[k]
    if pdf.get("doc_production"):
        d["doc_production"] = pdf["doc_production"][:14]   # major producers — keep the chart readable
    if docout is not None:
        d["doc_prod_current"] = docout
        d["doc_prod_current_label"] = f"DoC output ({latest_month})"
    if orb:
        hist = []
        if orb.get("y2025") is not None:
            hist.append({"m": "2025 avg", "v": orb["y2025"]})
        if orb.get("y2026") is not None:
            hist.append({"m": "2026 avg", "v": orb["y2026"]})
        if orb.get("prev") is not None:
            hist.append({"m": orb.get("prev_label", "prev"), "v": orb["prev"]})
        if orb.get("last") is not None:
            hist.append({"m": orb.get("last_label", "latest"), "v": orb["last"]})
        d["basket_history"] = hist

    # balance table
    rows = []
    if dg26 is not None:
        rows.append({"label": "World oil demand growth, 2026", "value": f"{dg26:+.1f} mb/d",
                     "mom": _rev_phrase(rev.get("demand_growth_2026")), "cls": "flat"})
    if dg27 is not None:
        rows.append({"label": "World oil demand growth, 2027", "value": f"{dg27:+.1f} mb/d",
                     "mom": _rev_phrase(rev.get("demand_growth_2027")), "cls": "flat"})
    if sg26 is not None:
        rows.append({"label": "Non-DoC liquids supply growth, 2026", "value": f"{sg26:+.1f} mb/d",
                     "mom": _rev_phrase(rev.get("supply_growth_2026")), "cls": "flat"})
    if c26 is not None:
        rows.append({"label": "Required DoC crude, 2026", "value": f"{c26:.1f} mb/d",
                     "mom": _rev_phrase(rev.get("call_2026")), "cls": "flat"})
    if c27 is not None:
        rows.append({"label": "Required DoC crude, 2027", "value": f"{c27:.1f} mb/d",
                     "mom": _rev_phrase(rev.get("call_2027")), "cls": "flat"})
    if docout is not None:
        rows.append({"label": f"DoC crude output, {latest_month} (secondary sources)",
                     "value": f"{docout:.1f} mb/d", "mom": "—", "cls": "flat"})
    if orb.get("last") is not None:
        rows.append({"label": f"OPEC Reference Basket, {latest_month} avg",
                     "value": f"${orb['last']:.2f}/b",
                     "mom": (f"+{orb['chg']:.2f}" if (orb.get('chg') or 0) >= 0 else f"{orb['chg']:.2f}"),
                     "cls": "flat"})
    d["balance_rows"] = rows
    d["balance_note"] = ("“DoC” = the OPEC and non-OPEC countries in the Declaration of Cooperation. "
                         "“Required DoC crude” (the call) = world demand minus non-DoC liquids and DoC NGLs; "
                         "the gap between that and actual DoC crude output is met by stock draws (call above "
                         "output) or builds (output above call). Demand and supply figures are OPEC's own.")

    # headline
    gap = (c26 - docout) if (c26 is not None and docout is not None) else None
    draw = gap is not None and gap > 0
    backw = " with the futures curves in steep backwardation" if pdf.get("backwardated") else ""
    parts = []
    if dg26 is not None and sg26 is not None:
        parts.append(f"OPEC's {month} report puts 2026 world oil demand growth at {dg26:+.1f} mb/d"
                     + (f" and 2027 at {dg27:+.1f} mb/d" if dg27 is not None else "")
                     + f", against non-DoC liquids supply growth of {sg26:+.1f} mb/d.")
    if c26 is not None and docout is not None:
        parts.append(f"Required DoC crude is put at {c26:.1f} mb/d for 2026"
                     + (f" and {c27:.1f} mb/d for 2027" if c27 is not None else "")
                     + f", versus the {docout:.1f} mb/d DoC producers actually pumped in {latest_month} — "
                     + ("implying a heavy draw on inventories." if draw else
                        "leaving the market broadly balanced."))
    if orb.get("last") is not None:
        parts.append(f"The OPEC Reference Basket averaged ${orb['last']:.2f}/b in {latest_month}{backw}"
                     + (", even as money managers cut net length." if pdf.get("specs_cut") else "."))
    d["headline"] = " ".join(parts)

    # implications (neutral observation)
    imp = []
    if draw:
        imp.append(f"Required DoC crude (~{c26:.1f} mb/d for 2026) sits well above current DoC output "
                   f"(~{docout:.1f} mb/d in {latest_month}), so on OPEC's own numbers the market is drawing "
                   f"on stocks — a backdrop {'consistent with the steep backwardation the report notes' if pdf.get('backwardated') else 'that tends to support prompt time-spreads'}.")
    elif gap is not None:
        imp.append(f"With current DoC output (~{docout:.1f} mb/d) close to the call (~{c26:.1f} mb/d), the "
                   f"report's balance looks broadly even — a setup that may be worth watching for either side to break.")
    if dg27 is not None and sg26 is not None and dg27 > sg26:
        imp.append(f"With 2027 demand growth at {dg27:+.1f} mb/d against non-DoC supply growth near {sg26:+.1f} mb/d, "
                   f"the call on DoC crude builds further next year; a widening call tends to show in the front of the curve first.")
    if orb.get("last") is not None:
        imp.append(f"The basket near ${orb['last']:.0f}" +
                   (f" (m/m {'+' if (orb.get('chg') or 0) >= 0 else ''}{orb.get('chg'):.2f})" if orb.get('chg') is not None else "") +
                   " and the report's stock picture point the same way; whether prompt spreads hold may hinge on how quickly DoC output moves.")
    if pdf.get("specs_cut"):
        imp.append("Money managers cut net length even as the report describes a tightening balance — a positioning-versus-fundamentals divergence that may be worth watching.")
    d["implications"] = imp[:4] or ["See the balance table and charts for this month's demand/supply picture."]

    # changes (m/m revisions + biggest production mover)
    ch = []
    if rev.get("demand_growth_2026") is not None or rev.get("demand_growth_2027") is not None:
        ch.append(f"2026 world demand growth {_rev_word(rev.get('demand_growth_2026'))}"
                  + (f"; 2027 {_rev_word(rev.get('demand_growth_2027'))}." if rev.get('demand_growth_2027') is not None else "."))
    if rev.get("supply_growth_2026") is not None:
        ch.append(f"Non-DoC supply growth {_rev_word(rev.get('supply_growth_2026'))} for 2026.")
    if rev.get("call_2026") is not None:
        ch.append(f"Required DoC crude for 2026 {_rev_word(rev.get('call_2026'))}"
                  + (f"; 2027 {_rev_word(rev.get('call_2027'))}." if rev.get('call_2027') is not None else "."))
    prod = pdf.get("doc_production") or []
    movers = [p for p in prod if p.get("mom") is not None]
    if movers:
        mv = max(movers, key=lambda p: abs(p["mom"]))
        if abs(mv["mom"]) >= 0.05:
            ch.append(f"Among producers, {mv['country']} showed the largest m/m move in {latest_month} "
                      f"({mv['mom']:+.2f} mb/d, secondary sources).")
    d["changes"] = ch or ["No material month-over-month revisions flagged this issue."]

    # EIA STEO cross-check — an independent forecaster to benchmark OPEC against (graceful).
    try:
        import eia
        comp = eia.build_comparison({
            "demand_growth_2026": dg26,
            "demand_level_2026": (appendix.get("levels") or {}).get("demand2026"),
            "nondoc_growth_2026": sg26,
            "orb_2026": orb.get("y2026"),
            "drawing": draw,
        })
        if comp.get("rows"):
            d["eia_comparison"] = comp["rows"]
            d["eia_note"] = comp.get("note", "")
            if comp.get("prices"):
                d["eia_prices"] = comp["prices"]
                d["eia_price_split"] = f"{year}-{_MONTHS.index(month) + 1:02d}"   # report month = actual/forecast divide
            if comp.get("callout"):
                d["implications"] = ([comp["callout"]] + d["implications"])[:4]
    except Exception:
        pass

    d["sources_note"] = (f"All figures from OPEC's Monthly Oil Market Report, {month} {year} "
                         "(Oil Market Highlights, Table 1-1, and Appendix Tables 11-1/11-2/11-4) "
                         "and the secondary-sources crude production table.")
    return d
